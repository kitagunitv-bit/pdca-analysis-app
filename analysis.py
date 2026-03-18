"""
PDCA分析エンジン — ふるさと納税返礼品 ABC分析
"""
import io
import warnings
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────
# スタイル定数
# ─────────────────────────────────────────────
FN = 'Arial'
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NAVY   = '1F4E79'
DKGREEN= '375623'
BROWN  = '7B3F00'
PURPLE = '762E8C'
DKRED  = 'C00000'

RANK_A = ('C6EFCE', '276221')
RANK_B = ('FFEB9C', '9C6500')
RANK_C = ('FFC7CE', '9C0006')
RANK_MAP = {'A': RANK_A, 'B': RANK_B, 'C': RANK_C}

YEAR_COLORS = {'1年生': '4472C4', '2年生': 'ED7D31', '3年生': '70AD47'}

def hdr(ws, r, c, val, bg=NAVY, fg='FFFFFF', bold=True, sz=11):
    cell = ws.cell(r, c, val)
    cell.font = Font(name=FN, bold=bold, color=fg, size=sz)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
    return cell

def num_cell(ws, r, c, val, fmt='#,##0'):
    cell = ws.cell(r, c, val)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = BORDER
    cell.font = Font(name=FN, size=10)
    return cell

def pct_cell(ws, r, c, val):
    cell = ws.cell(r, c, val)
    cell.number_format = '0.0%'
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = BORDER
    cell.font = Font(name=FN, size=10)
    return cell

def txt_cell(ws, r, c, val, align='center'):
    cell = ws.cell(r, c, val)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER
    cell.font = Font(name=FN, size=10)
    return cell

def rank_cell(ws, r, c, rank):
    bg, fg = RANK_MAP.get(rank, RANK_B)
    cell = ws.cell(r, c, rank)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.font = Font(name=FN, bold=True, color=fg, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER

def title_row(ws, r, max_col, text, bg=NAVY, sz=14, ht=32):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    c = ws.cell(r, 1, text)
    c.font = Font(name=FN, bold=True, color='FFFFFF', size=sz)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = ht

def set_col_widths(ws, widths, start=1):
    for i, w in enumerate(widths, start):
        ws.column_dimensions[get_column_letter(i)].width = w

def apply_border(ws, r1, r2, c1, c2):
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = BORDER

def assign_abc(series, top=0.2, bottom=0.2):
    n = len(series)
    rank = series.rank(method='first', ascending=False)
    a_thresh = n * top
    c_thresh = n * (1 - bottom)
    result = pd.Series('B', index=series.index)
    result[rank <= a_thresh] = 'A'
    result[rank > c_thresh] = 'C'
    return result

# ─────────────────────────────────────────────
# データ読み込み
# ─────────────────────────────────────────────
def load_data(file_obj):
    """アップロードされたExcelファイルを読み込んでDataFrameを返す"""
    df_raw = pd.read_excel(file_obj, sheet_name=None, header=None)

    # シートを自動検出: 行数が最も多いシートを使用
    target_sheet = None
    max_rows = 0
    for sheet_name, df in df_raw.items():
        if len(df) > max_rows:
            max_rows = len(df)
            target_sheet = sheet_name

    df_raw = pd.read_excel(file_obj, sheet_name=target_sheet, header=None)

    # ヘッダー検出（「商品コード」が含まれる行を探す）
    header_row = None
    for i, row in df_raw.iterrows():
        if any('商品コード' in str(v) for v in row.values):
            header_row = i
            break

    if header_row is None:
        # デフォルト: 行1がヘッダー
        header_row = 1

    data = df_raw.iloc[header_row + 2:].copy()
    data.columns = range(len(data.columns))

    col_map = {
        0: '商品コード', 1: '返礼品名', 2: 'OG', 3: 'OGファミリー',
        4: 'カテゴリ', 5: '分類', 6: '販売年数',
        7: '寄付額', 8: '返礼額', 9: '商品原価', 10: '単位粗利益',
    }
    months = ['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
    for i, m in enumerate(months):
        base = 11 + i * 3
        col_map[base]   = f'{m}受注'
        col_map[base+1] = f'{m}売上'
        col_map[base+2] = f'{m}粗利'
    col_map[47] = '合計受注件数'
    col_map[48] = '合計売上金額'
    col_map[49] = '合計粗利益'

    data = data.rename(columns=col_map)

    num_cols = ['寄付額','返礼額','商品原価','単位粗利益','合計受注件数','合計売上金額','合計粗利益']
    for m in months:
        num_cols += [f'{m}受注', f'{m}売上', f'{m}粗利']
    for c in num_cols:
        if c in data.columns:
            data[c] = pd.to_numeric(data[c], errors='coerce').fillna(0)

    data = data.dropna(subset=['商品コード']).copy()
    data['販売年数'] = data['販売年数'].astype(str).str.strip()

    return data, months, sheet_name

# ─────────────────────────────────────────────
# 分析処理
# ─────────────────────────────────────────────
def compute_analysis(data, months):
    data = data.copy()

    data['粗利率'] = data.apply(
        lambda r: r['合計粗利益'] / r['合計売上金額'] if r['合計売上金額'] > 0 else 0, axis=1)

    data['受注ランク'] = assign_abc(data['合計受注件数'])
    data['売上ランク'] = assign_abc(data['合計売上金額'])
    data['粗利ランク'] = assign_abc(data['合計粗利益'])

    # 累積比率（各指標の降順）
    for metric, cum_col in [('合計受注件数','受注累積比率'),
                             ('合計売上金額','売上累積比率'),
                             ('合計粗利益', '粗利累積比率')]:
        s = data.sort_values(metric, ascending=False)[metric]
        total = s.sum()
        data.loc[s.index, cum_col] = s.cumsum() / total if total > 0 else 0

    # 4象限
    median_sales = data['合計売上金額'].median()
    def classify_q(row):
        hi = row['合計売上金額'] >= median_sales
        gp = row['粗利率']
        if gp >= 0.20 and hi:  return 'Q1: スター'
        elif gp >= 0.20:       return 'Q2: 高収益'
        elif gp <= 0.10 and hi: return 'Q3: 量販型'
        else:                  return 'Q4: 要改善'
    data['象限分類'] = data.apply(classify_q, axis=1)

    TOTAL_ORDERS = int(data['合計受注件数'].sum())
    TOTAL_SALES  = int(data['合計売上金額'].sum())
    TOTAL_GP     = int(data['合計粗利益'].sum())
    TOTAL_ITEMS  = len(data)
    GP_RATE      = TOTAL_GP / TOTAL_SALES if TOTAL_SALES else 0

    monthly_orders = [data[f'{m}受注'].sum() for m in months]
    monthly_sales  = [data[f'{m}売上'].sum() for m in months]
    monthly_gp     = [data[f'{m}粗利'].sum() for m in months]

    stats = {
        'TOTAL_ORDERS': TOTAL_ORDERS,
        'TOTAL_SALES':  TOTAL_SALES,
        'TOTAL_GP':     TOTAL_GP,
        'TOTAL_ITEMS':  TOTAL_ITEMS,
        'GP_RATE':      GP_RATE,
        'median_sales': median_sales,
        'monthly_orders': monthly_orders,
        'monthly_sales':  monthly_sales,
        'monthly_gp':     monthly_gp,
    }
    return data, stats

# ─────────────────────────────────────────────
# Excel 出力
# ─────────────────────────────────────────────
def build_excel(data, stats, months, title_prefix='泉佐野市'):
    wb = Workbook()

    TOTAL_ORDERS   = stats['TOTAL_ORDERS']
    TOTAL_SALES    = stats['TOTAL_SALES']
    TOTAL_GP       = stats['TOTAL_GP']
    TOTAL_ITEMS    = stats['TOTAL_ITEMS']
    GP_RATE        = stats['GP_RATE']
    median_sales   = stats['median_sales']
    monthly_orders = stats['monthly_orders']
    monthly_sales  = stats['monthly_sales']
    monthly_gp     = stats['monthly_gp']

    # ══════════════════════════════════════════
    # Sheet 1: ダッシュボード
    # ══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'ダッシュボード'
    ws1.sheet_properties.tabColor = NAVY

    title_row(ws1, 1, 14, f'{title_prefix} ふるさと納税 返礼品 PDCA分析 ダッシュボード', bg=NAVY, sz=16, ht=40)

    kpi_data = [
        ('総商品数',    str(TOTAL_ITEMS),       '商品', NAVY),
        ('年間受注件数', f'{TOTAL_ORDERS:,}',    '件',   '2E75B6'),
        ('年間売上金額', f'¥{TOTAL_SALES:,}',    '',     DKGREEN),
        ('年間粗利益',  f'¥{TOTAL_GP:,}',        '',     BROWN),
        ('粗利率',      f'{GP_RATE*100:.1f}%',   '',     DKRED),
    ]
    for i, (label, val, unit, bg) in enumerate(kpi_data):
        col = i * 3 + 1
        ws1.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        c_label = ws1.cell(3, col, label)
        c_label.font = Font(name=FN, bold=True, color='FFFFFF', size=10)
        c_label.fill = PatternFill('solid', start_color=bg)
        c_label.alignment = Alignment(horizontal='center', vertical='center')

        ws1.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        c_val = ws1.cell(4, col, f'{val}{unit}')
        c_val.font = Font(name=FN, bold=True, color=bg, size=15)
        c_val.alignment = Alignment(horizontal='center', vertical='center')
        c_val.fill = PatternFill('solid', start_color='F2F2F2')
    ws1.row_dimensions[3].height = 22
    ws1.row_dimensions[4].height = 36

    tbl_start = 6
    hdr(ws1, tbl_start, 1, '月', bg='404040', sz=10)
    for i, m in enumerate(months, 2):
        hdr(ws1, tbl_start, i, m, bg='404040', sz=10)
    hdr(ws1, tbl_start, 14, '合計', bg='404040', sz=10)

    for ri, (lbl, vals, total) in enumerate([
        ('受注件数', monthly_orders, TOTAL_ORDERS),
        ('売上金額', monthly_sales,  TOTAL_SALES),
        ('粗利益',  monthly_gp,     TOTAL_GP),
    ]):
        r = tbl_start + 1 + ri
        txt_cell(ws1, r, 1, lbl)
        ws1.cell(r, 1).font = Font(name=FN, bold=True, size=10)
        for ci, v in enumerate(vals, 2):
            num_cell(ws1, r, ci, int(v))
        num_cell(ws1, r, 14, int(total))
        ws1.cell(r, 14).font = Font(name=FN, bold=True, size=10)

    ws1.column_dimensions['A'].width = 12
    for i in range(2, 15):
        ws1.column_dimensions[get_column_letter(i)].width = 12

    chart1 = BarChart()
    chart1.type = 'col'
    chart1.title = '月別売上金額推移'
    chart1.y_axis.title = '売上金額(円)'
    chart1.x_axis.title = '月'
    chart1.style = 10
    chart1.width = 28
    chart1.height = 14
    cats = Reference(ws1, min_col=2, max_col=13, min_row=tbl_start)
    d_sales = Reference(ws1, min_col=2, max_col=13, min_row=tbl_start+2)
    chart1.add_data(d_sales, from_rows=True, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.series[0].title = SeriesLabel(v='売上金額')
    chart1.series[0].graphicalProperties.solidFill = '2E75B6'
    chart1.legend = None
    ws1.add_chart(chart1, 'A11')

    chart2 = BarChart()
    chart2.type = 'col'
    chart2.title = '月別受注件数推移'
    chart2.y_axis.title = '受注件数'
    chart2.style = 10
    chart2.width = 28
    chart2.height = 14
    d_orders = Reference(ws1, min_col=2, max_col=13, min_row=tbl_start+1)
    chart2.add_data(d_orders, from_rows=True, titles_from_data=False)
    chart2.set_categories(cats)
    chart2.series[0].title = SeriesLabel(v='受注件数')
    chart2.series[0].graphicalProperties.solidFill = 'ED7D31'
    chart2.legend = None
    ws1.add_chart(chart2, 'A28')

    sum_start = 45
    title_row(ws1, sum_start, 14, '全体サマリー（カテゴリ別）', bg='404040', sz=12, ht=24)
    sum_hdrs = ['カテゴリ','商品数','受注件数','売上金額','粗利益','粗利率','売上構成比']
    for i, h in enumerate(sum_hdrs, 1):
        hdr(ws1, sum_start+1, i, h, bg='404040', sz=10)

    cat_agg = data.groupby('カテゴリ').agg(
        商品数=('商品コード','count'),
        受注件数=('合計受注件数','sum'),
        売上金額=('合計売上金額','sum'),
        粗利益=('合計粗利益','sum')
    ).sort_values('売上金額', ascending=False).reset_index()

    for idx, row in cat_agg.iterrows():
        r = sum_start + 2 + idx
        txt_cell(ws1, r, 1, row['カテゴリ'])
        num_cell(ws1, r, 2, int(row['商品数']))
        num_cell(ws1, r, 3, int(row['受注件数']))
        num_cell(ws1, r, 4, int(row['売上金額']))
        num_cell(ws1, r, 5, int(row['粗利益']))
        pct_cell(ws1, r, 6, row['粗利益']/row['売上金額'] if row['売上金額']>0 else 0)
        pct_cell(ws1, r, 7, row['売上金額']/TOTAL_SALES if TOTAL_SALES>0 else 0)

    tr1 = sum_start + 2 + len(cat_agg)
    for ci in range(1, 8):
        ws1.cell(tr1, ci).fill = PatternFill('solid', start_color='D9E1F2')
        ws1.cell(tr1, ci).border = BORDER
        ws1.cell(tr1, ci).font = Font(name=FN, bold=True, size=10)
    txt_cell(ws1, tr1, 1, '合計')
    num_cell(ws1, tr1, 2, TOTAL_ITEMS)
    num_cell(ws1, tr1, 3, TOTAL_ORDERS)
    num_cell(ws1, tr1, 4, TOTAL_SALES)
    num_cell(ws1, tr1, 5, TOTAL_GP)
    pct_cell(ws1, tr1, 6, GP_RATE)
    pct_cell(ws1, tr1, 7, 1.0)
    for ci in range(1, 8):
        ws1.cell(tr1, ci).font = Font(name=FN, bold=True, size=10)
    apply_border(ws1, sum_start+1, tr1, 1, 7)
    ws1.auto_filter.ref = f'A{sum_start+1}:G{tr1}'

    # ══════════════════════════════════════════
    # Sheet 2: 商品データ
    # ══════════════════════════════════════════
    ws2 = wb.create_sheet('商品データ')
    ws2.sheet_properties.tabColor = '2E75B6'

    s2_headers = ['No','商品コード','返礼品名','OG','OGファミリー','カテゴリ','分類','販売年数',
                  '寄付額','返礼額','商品原価',
                  '合計受注件数','受注累積比率','受注ランク',
                  '合計売上金額','売上累積比率','売上ランク',
                  '合計粗利益','粗利累積比率','粗利ランク',
                  '粗利率','象限分類']
    s2_widths = [5,13,44,12,15,10,10,9,10,10,10,12,12,10,14,12,10,12,12,10,10,16]

    title_row(ws2, 1, len(s2_headers), '商品データ一覧（ABC分析 + 4象限分類）', bg='2E75B6', ht=30)
    for i, (h, w) in enumerate(zip(s2_headers, s2_widths), 1):
        hdr(ws2, 2, i, h, bg='2E75B6')
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 30

    sorted_d = data.sort_values('合計売上金額', ascending=False).reset_index(drop=True)

    q_colors = {
        'Q1: スター': RANK_A,
        'Q2: 高収益': ('BDD7EE', '1F4E79'),
        'Q3: 量販型': RANK_C,
        'Q4: 要改善': RANK_B,
    }

    for idx, row in sorted_d.iterrows():
        r = idx + 3
        txt_cell(ws2, r, 1, idx+1)
        txt_cell(ws2, r, 2, row['商品コード'])
        txt_cell(ws2, r, 3, row['返礼品名'], align='left')
        txt_cell(ws2, r, 4, row['OG'])
        txt_cell(ws2, r, 5, row['OGファミリー'])
        txt_cell(ws2, r, 6, row['カテゴリ'])
        txt_cell(ws2, r, 7, row['分類'])
        txt_cell(ws2, r, 8, row['販売年数'])
        num_cell(ws2, r, 9, int(row['寄付額']))
        num_cell(ws2, r, 10, int(row['返礼額']))
        num_cell(ws2, r, 11, int(row['商品原価']))
        num_cell(ws2, r, 12, int(row['合計受注件数']))
        pct_cell(ws2, r, 13, row['受注累積比率'])
        rank_cell(ws2, r, 14, row['受注ランク'])
        num_cell(ws2, r, 15, int(row['合計売上金額']))
        pct_cell(ws2, r, 16, row['売上累積比率'])
        rank_cell(ws2, r, 17, row['売上ランク'])
        num_cell(ws2, r, 18, int(row['合計粗利益']))
        pct_cell(ws2, r, 19, row['粗利累積比率'])
        rank_cell(ws2, r, 20, row['粗利ランク'])
        pct_cell(ws2, r, 21, row['粗利率'])
        qbg, qfg = q_colors.get(row['象限分類'], RANK_B)
        c22 = ws2.cell(r, 22, row['象限分類'])
        c22.fill = PatternFill('solid', start_color=qbg)
        c22.font = Font(name=FN, bold=True, color=qfg, size=10)
        c22.alignment = Alignment(horizontal='center', vertical='center')
        c22.border = BORDER
        ws2.row_dimensions[r].height = 18

    ws2.auto_filter.ref = f'A2:V{len(sorted_d)+2}'
    ws2.freeze_panes = 'A3'

    # ══════════════════════════════════════════
    # Sheet 3: OG別ABC分析
    # ══════════════════════════════════════════
    ws3 = wb.create_sheet('OG別ABC分析')
    ws3.sheet_properties.tabColor = DKGREEN

    og_agg = data.groupby('OG').agg(
        商品数=('商品コード','count'),
        合計受注件数=('合計受注件数','sum'),
        合計売上金額=('合計売上金額','sum'),
        合計粗利益=('合計粗利益','sum')
    ).reset_index()
    og_agg['粗利率'] = og_agg.apply(
        lambda r: r['合計粗利益']/r['合計売上金額'] if r['合計売上金額']>0 else 0, axis=1)
    og_agg['受注ランク'] = assign_abc(og_agg['合計受注件数'])
    og_agg['売上ランク'] = assign_abc(og_agg['合計売上金額'])
    og_agg['粗利ランク'] = assign_abc(og_agg['合計粗利益'])
    og_agg['受注シェア'] = og_agg['合計受注件数'] / TOTAL_ORDERS
    og_agg['売上シェア'] = og_agg['合計売上金額'] / TOTAL_SALES
    og_agg['粗利シェア'] = og_agg['合計粗利益']  / TOTAL_GP

    for metric, cum_col in [('合計受注件数','受注累積比率'),('合計売上金額','売上累積比率'),('合計粗利益','粗利累積比率')]:
        s = og_agg.sort_values(metric, ascending=False)[metric]
        og_agg.loc[s.index, cum_col] = s.cumsum() / s.sum()

    og_agg = og_agg.sort_values('合計売上金額', ascending=False).reset_index(drop=True)

    title_row(ws3, 1, 16, 'OG（オリジン）別 ABC分析 ― 受注件数・売上金額・粗利益 横並び比較', bg=DKGREEN, ht=32)
    ws3.merge_cells('A2:P2')
    ws3.cell(2,1,'【ABCランク基準】 A: 上位20% / B: 中間60% / C: 下位20%  |  累積比率は各指標降順で算出').font = Font(name=FN,size=10,color='595959')
    ws3.row_dimensions[2].height = 18

    og_hdrs = ['No','OG','商品数',
               '受注件数','受注シェア','受注累積比率','受注ランク',
               '売上金額','売上シェア','売上累積比率','売上ランク',
               '粗利益','粗利シェア','粗利累積比率','粗利ランク',
               '粗利率']
    og_ws   = [5,15,7, 12,10,11,9, 14,10,11,9, 12,10,11,9, 10]
    for i, (h, w) in enumerate(zip(og_hdrs, og_ws), 1):
        hdr(ws3, 3, i, h, bg=DKGREEN)
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[3].height = 30

    for idx, row in og_agg.iterrows():
        r = idx + 4
        txt_cell(ws3, r, 1, idx+1)
        txt_cell(ws3, r, 2, row['OG'])
        num_cell(ws3, r, 3, int(row['商品数']))
        num_cell(ws3, r, 4, int(row['合計受注件数']))
        pct_cell(ws3, r, 5, row['受注シェア'])
        pct_cell(ws3, r, 6, row['受注累積比率'])
        rank_cell(ws3, r, 7, row['受注ランク'])
        num_cell(ws3, r, 8, int(row['合計売上金額']))
        pct_cell(ws3, r, 9, row['売上シェア'])
        pct_cell(ws3, r, 10, row['売上累積比率'])
        rank_cell(ws3, r, 11, row['売上ランク'])
        num_cell(ws3, r, 12, int(row['合計粗利益']))
        pct_cell(ws3, r, 13, row['粗利シェア'])
        pct_cell(ws3, r, 14, row['粗利累積比率'])
        rank_cell(ws3, r, 15, row['粗利ランク'])
        pct_cell(ws3, r, 16, row['粗利率'])
        ws3.row_dimensions[r].height = 18

    tr3 = len(og_agg) + 4
    for ci in range(1, 17):
        ws3.cell(tr3, ci).fill = PatternFill('solid', start_color='D9E1F2')
        ws3.cell(tr3, ci).border = BORDER
        ws3.cell(tr3, ci).font = Font(name=FN, bold=True, size=10)
    txt_cell(ws3, tr3, 2, f'{len(og_agg)} OG')
    num_cell(ws3, tr3, 3, TOTAL_ITEMS)
    num_cell(ws3, tr3, 4, TOTAL_ORDERS)
    pct_cell(ws3, tr3, 5, 1.0)
    pct_cell(ws3, tr3, 6, 1.0)
    num_cell(ws3, tr3, 8, TOTAL_SALES)
    pct_cell(ws3, tr3, 9, 1.0)
    pct_cell(ws3, tr3, 10, 1.0)
    num_cell(ws3, tr3, 12, TOTAL_GP)
    pct_cell(ws3, tr3, 13, 1.0)
    pct_cell(ws3, tr3, 14, 1.0)
    pct_cell(ws3, tr3, 16, GP_RATE)
    for ci in range(1, 17):
        ws3.cell(tr3, ci).font = Font(name=FN, bold=True, size=10)
    ws3.auto_filter.ref = f'A3:P{tr3}'
    ws3.freeze_panes = 'A4'

    # ══════════════════════════════════════════
    # Sheet 4: 販売年数別ABC分析
    # ══════════════════════════════════════════
    ws4 = wb.create_sheet('販売年数別ABC分析')
    ws4.sheet_properties.tabColor = BROWN

    title_row(ws4, 1, 14, '販売年数別 ABC分析 ― 各年数グループ内でランク付け', bg=BROWN, ht=32)
    ws4.merge_cells('A2:N2')
    ws4.cell(2,1,'【ABCランク基準】各販売年数グループ内で独立にランク: A=上位20% / B=中間60% / C=下位20%').font = Font(name=FN,size=10,color='595959')

    y4_hdrs = ['No','商品コード','返礼品名','OG','OGファミリー','カテゴリ','販売年数',
               '受注件数','受注ランク','売上金額','売上ランク','粗利益','粗利ランク','粗利率']
    y4_ws   = [5,13,44,12,15,10,9, 12,10,14,10,12,10,10]
    for i, (h, w) in enumerate(zip(y4_hdrs, y4_ws), 1):
        hdr(ws4, 3, i, h, bg=BROWN)
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.row_dimensions[3].height = 30

    cr4 = 4
    for year in ['1年生','2年生','3年生']:
        yd = data[data['販売年数'] == year].copy()
        if yd.empty:
            continue
        yd['yr_受注ランク'] = assign_abc(yd['合計受注件数'])
        yd['yr_売上ランク'] = assign_abc(yd['合計売上金額'])
        yd['yr_粗利ランク'] = assign_abc(yd['合計粗利益'])
        yd = yd.sort_values('合計売上金額', ascending=False).reset_index(drop=True)

        bg_y = YEAR_COLORS.get(year, NAVY)
        ws4.merge_cells(start_row=cr4, start_column=1, end_row=cr4, end_column=len(y4_hdrs))
        gc = ws4.cell(cr4, 1, f'▶ {year}（{len(yd)}商品）')
        gc.font = Font(name=FN, bold=True, color='FFFFFF', size=11)
        gc.fill = PatternFill('solid', start_color=bg_y)
        gc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws4.row_dimensions[cr4].height = 24
        cr4 += 1

        for idx, row in yd.iterrows():
            r = cr4
            txt_cell(ws4, r, 1, idx+1)
            txt_cell(ws4, r, 2, row['商品コード'])
            txt_cell(ws4, r, 3, row['返礼品名'], align='left')
            txt_cell(ws4, r, 4, row['OG'])
            txt_cell(ws4, r, 5, row['OGファミリー'])
            txt_cell(ws4, r, 6, row['カテゴリ'])
            yc = txt_cell(ws4, r, 7, row['販売年数'])
            yc.font = Font(name=FN, bold=True, color=bg_y, size=10)
            num_cell(ws4, r, 8, int(row['合計受注件数']))
            rank_cell(ws4, r, 9, row['yr_受注ランク'])
            num_cell(ws4, r, 10, int(row['合計売上金額']))
            rank_cell(ws4, r, 11, row['yr_売上ランク'])
            num_cell(ws4, r, 12, int(row['合計粗利益']))
            rank_cell(ws4, r, 13, row['yr_粗利ランク'])
            pct_cell(ws4, r, 14, row['粗利率'])
            ws4.row_dimensions[r].height = 18
            cr4 += 1

        ws4.merge_cells(start_row=cr4, start_column=1, end_row=cr4, end_column=7)
        sc = ws4.cell(cr4, 1, f'{year} 小計')
        sc.font = Font(name=FN, bold=True, size=10)
        sc.fill = PatternFill('solid', start_color='D9E1F2')
        sc.alignment = Alignment(horizontal='right', vertical='center')
        for ci in range(1, 15):
            ws4.cell(cr4, ci).fill = PatternFill('solid', start_color='D9E1F2')
            ws4.cell(cr4, ci).border = BORDER
        num_cell(ws4, cr4, 8, int(yd['合計受注件数'].sum()))
        ws4.cell(cr4, 8).font = Font(name=FN, bold=True)
        num_cell(ws4, cr4, 10, int(yd['合計売上金額'].sum()))
        ws4.cell(cr4, 10).font = Font(name=FN, bold=True)
        num_cell(ws4, cr4, 12, int(yd['合計粗利益'].sum()))
        ws4.cell(cr4, 12).font = Font(name=FN, bold=True)
        s_gp = yd['合計粗利益'].sum()
        s_sl = yd['合計売上金額'].sum()
        pct_cell(ws4, cr4, 14, s_gp/s_sl if s_sl>0 else 0)
        ws4.cell(cr4, 14).font = Font(name=FN, bold=True)
        ws4.row_dimensions[cr4].height = 22
        cr4 += 2

    ws4.auto_filter.ref = f'A3:N{cr4-1}'
    ws4.freeze_panes = 'A4'

    # ══════════════════════════════════════════
    # Sheet 5: 販売年数別円グラフ
    # ══════════════════════════════════════════
    ws5 = wb.create_sheet('販売年数別円グラフ')
    ws5.sheet_properties.tabColor = PURPLE

    title_row(ws5, 1, 8, '販売年数別 構成比グラフ', bg=PURPLE, ht=30)

    yr_summ = data.groupby('販売年数').agg(
        商品数=('商品コード','count'),
        受注件数=('合計受注件数','sum'),
        売上金額=('合計売上金額','sum'),
        粗利益=('合計粗利益','sum')
    ).reindex(['1年生','2年生','3年生']).reset_index()
    yr_summ['粗利率'] = yr_summ['粗利益'] / yr_summ['売上金額']

    for i, h in enumerate(['販売年数','商品数','受注件数','売上金額','粗利益','粗利率'], 1):
        hdr(ws5, 3, i, h, bg=PURPLE)
    ws5.row_dimensions[3].height = 26

    for idx, row in yr_summ.iterrows():
        r = idx + 4
        bg_y = YEAR_COLORS.get(str(row['販売年数']), 'FFFFFF')
        c1 = ws5.cell(r, 1, row['販売年数'])
        c1.fill = PatternFill('solid', start_color=bg_y)
        c1.font = Font(name=FN, bold=True, color='FFFFFF')
        c1.alignment = Alignment(horizontal='center', vertical='center')
        c1.border = BORDER
        num_cell(ws5, r, 2, int(row['商品数']))
        num_cell(ws5, r, 3, int(row['受注件数']))
        num_cell(ws5, r, 4, int(row['売上金額']))
        num_cell(ws5, r, 5, int(row['粗利益']))
        pct_cell(ws5, r, 6, row['粗利率'])

    tr5 = len(yr_summ) + 4
    for ci in range(1, 7):
        ws5.cell(tr5, ci).fill = PatternFill('solid', start_color='D9E1F2')
        ws5.cell(tr5, ci).border = BORDER
    txt_cell(ws5, tr5, 1, '合計')
    for c_i, v in enumerate([TOTAL_ITEMS,TOTAL_ORDERS,TOTAL_SALES,TOTAL_GP], 2):
        num_cell(ws5, tr5, c_i, v)
        ws5.cell(tr5, c_i).font = Font(name=FN, bold=True)
    pct_cell(ws5, tr5, 6, GP_RATE)
    ws5.cell(tr5, 6).font = Font(name=FN, bold=True)
    apply_border(ws5, 3, tr5, 1, 6)
    set_col_widths(ws5, [12, 8, 14, 16, 14, 10])

    for title_sfx, dcol, anchor in [('受注件数',3,'A9'),('売上金額',4,'G9'),('粗利益',5,'A27')]:
        pie = PieChart()
        pie.title = f'販売年数別 {title_sfx} 構成比'
        pie.style = 10
        pie.width = 16
        pie.height = 14
        labels = Reference(ws5, min_col=1, min_row=4, max_row=4+len(yr_summ)-1)
        dref = Reference(ws5, min_col=dcol, min_row=3, max_row=3+len(yr_summ))
        pie.add_data(dref, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = False
        ws5.add_chart(pie, anchor)

    # ══════════════════════════════════════════
    # Sheet 6: 4象限スター分析
    # ══════════════════════════════════════════
    ws6 = wb.create_sheet('4象限スター分析')
    ws6.sheet_properties.tabColor = DKRED

    title_row(ws6, 1, 10, '4象限スター分析 ― 売上金額 × 粗利率', bg=DKRED, ht=32)

    q_defs = [
        ('Q1: スター',  '高売上 × 粗利率≥20%', 'C6EFCE', '276221'),
        ('Q2: 高収益',  '低売上 × 粗利率≥20%', 'BDD7EE', '1F4E79'),
        ('Q3: 量販型',  '高売上 × 粗利率≤10%', 'FFC7CE', '9C0006'),
        ('Q4: 要改善',  '低売上 × 粗利率≤10%', 'FFEB9C', '9C6500'),
    ]
    for i, (ql, qd, qbg, qfg) in enumerate(q_defs):
        col = i * 3 + 1
        ws6.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        c = ws6.cell(2, col, f'{ql}: {qd}')
        c.fill = PatternFill('solid', start_color=qbg)
        c.font = Font(name=FN, bold=True, color=qfg, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws6.row_dimensions[2].height = 20

    ws6.merge_cells('A3:J3')
    ws6.cell(3,1,f'※売上金額の高低基準（中央値）: ¥{int(median_sales):,}').font = Font(name=FN,size=10,color='595959',italic=True)

    m6_hdrs = ['No','商品コード','返礼品名','OG','カテゴリ','販売年数',
               '合計売上金額','合計粗利益','粗利率','象限分類']
    m6_ws   = [5,13,44,12,10,9,14,12,10,16]
    for i, (h, w) in enumerate(zip(m6_hdrs, m6_ws), 1):
        hdr(ws6, 4, i, h, bg=DKRED)
        ws6.column_dimensions[get_column_letter(i)].width = w
    ws6.row_dimensions[4].height = 28

    cr6 = 5
    for quad, qbg, qfg in [
        ('Q1: スター', 'C6EFCE', '276221'),
        ('Q2: 高収益', 'BDD7EE', '1F4E79'),
        ('Q3: 量販型', 'FFC7CE', '9C0006'),
        ('Q4: 要改善', 'FFEB9C', '9C6500'),
    ]:
        qd = data[data['象限分類']==quad].sort_values('合計売上金額', ascending=False).reset_index(drop=True)
        if qd.empty:
            continue
        ws6.merge_cells(start_row=cr6, start_column=1, end_row=cr6, end_column=len(m6_hdrs))
        gc6 = ws6.cell(cr6, 1, f'▶ {quad}（{len(qd)}商品）')
        gc6.font = Font(name=FN, bold=True, color=qfg, size=11)
        gc6.fill = PatternFill('solid', start_color=qbg)
        gc6.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws6.row_dimensions[cr6].height = 24
        cr6 += 1
        for idx, row in qd.iterrows():
            r = cr6
            txt_cell(ws6, r, 1, idx+1)
            txt_cell(ws6, r, 2, row['商品コード'])
            txt_cell(ws6, r, 3, row['返礼品名'], align='left')
            txt_cell(ws6, r, 4, row['OG'])
            txt_cell(ws6, r, 5, row['カテゴリ'])
            txt_cell(ws6, r, 6, row['販売年数'])
            num_cell(ws6, r, 7, int(row['合計売上金額']))
            num_cell(ws6, r, 8, int(row['合計粗利益']))
            pct_cell(ws6, r, 9, row['粗利率'])
            c10 = ws6.cell(r, 10, quad)
            c10.fill = PatternFill('solid', start_color=qbg)
            c10.font = Font(name=FN, bold=True, color=qfg, size=10)
            c10.alignment = Alignment(horizontal='center', vertical='center')
            c10.border = BORDER
            ws6.row_dimensions[r].height = 18
            cr6 += 1
        cr6 += 1

    ws6.auto_filter.ref = f'A4:J{cr6-1}'
    ws6.freeze_panes = 'A5'

    # ──── BytesIO に書き出して返す ────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
